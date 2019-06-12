from openpyxl import load_workbook


class ExcelImporter:

    def __init__(self):
        self.__relationships = dict()

    def set_relationships(self, ws, statistic):
        relationships = self.__relationships
        relationships[statistic] = dict()
        for row in ws.iter_rows(values_only=True):
            if row[0] in relationships[statistic]:
                for ind, i in enumerate(row):
                    if ind > 0 and i is not None and i not in relationships[statistic][row[0]]:
                        relationships[statistic][row[0]].append(i)
            else:
                relationships[statistic][row[0]] = [i for ind, i in enumerate(row) if ind > 0 and i is not None]

    def get_relationships(self):
        return self.__relationships

    def build(self, path):
        wb = load_workbook(path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            self.set_relationships(ws=ws, statistic=sheet)
