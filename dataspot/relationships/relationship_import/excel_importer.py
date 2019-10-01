from openpyxl import load_workbook


class ExcelImporter:
    """

    """

    def __init__(self):
        self.__relationships = None

    def set_relationships(self, ws):
        """

        :param ws:
        :type ws:
        """
        relationships = dict()
        for row in ws.iter_rows(values_only=True):
            if row[0].lower() in relationships:
                for ind, i in enumerate(row):
                    if ind > 0 and i is not None and i.lower() not in relationships[row[0].lower()]:
                        relationships[row[0].lower()].append(i.lower())
            else:
                relationships[row[0].lower()] = list()
                for ind, source in enumerate(row):
                    if ind > 0 and source is not None:
                        relationships[row[0].lower()].append(source.lower())

        self.__relationships = relationships

    def get_relationships(self):
        """
        :return:
        :rtype:
        """
        return self.__relationships

    def build(self, path):
        """

        :param path:
        :type path:
        """
        wb = load_workbook(path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            self.set_relationships(ws=ws)


excel = ExcelImporter()
excel.build(path='/Users/patrickdehoon/Projecten/prive/dataspot/clan/excel/clan-sources.xlsx')
clan_rel = excel.get_relationships()

excel = ExcelImporter()
excel.build(path='/Users/patrickdehoon/Projecten/prive/dataspot/clan/excel/mstr overzicht.xlsx')
mstr_rel = excel.get_relationships()

new_rel = {**clan_rel, **mstr_rel}
print(new_rel)